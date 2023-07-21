import os

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase

from HHPG.application.HaushaltsplanExcelGenerator import HaushaltsplanExcelGenerator
from HHPG.domain.entity.aufwand import Aufwand
from HHPG.domain.entity.haushaltsplan import Haushaltsplan
from HHPG.domain.entity.haushaltsposten import Haushaltsposten
from HHPG.domain.entity.projekt import Projekt
from HHPG.infrastruktur.haushaltsplan_repository import HaushaltsplanRepository


class HaushaltsplanExcelGeneratorTestCase(TestCase):
    def setUp(self):
        # Create a test user (replace with actual user creation if needed)
        self.user = User.objects.create_user(username='testuser', password='testpassword')

        # Create a Haushaltsplan instance
        self.haushaltsplan = Haushaltsplan.objects.create(
            plan_name='Test Plan',
            standort='Heilbronn',
            startjahr=2023,
            author=self.user,
            studierendenzahl=1000,
        )

        # Create a Haushaltsposten instance associated with the Haushaltsplan
        self.haushaltsposten = Haushaltsposten.objects.create(
            haushaltsplan=self.haushaltsplan,
            posten_name='Test Posten',
        )

        # Create a Projekt instance associated with the Haushaltsposten
        self.projekt = Projekt.objects.create(
            projekt_name='Test Projekt',
            haushaltsposten=self.haushaltsposten,
        )

        # Create an Aufwand instance associated with the Projekt
        self.aufwand = Aufwand.objects.create(
            einnahmen=5000,
            ausgaben=3000,
            projekt=self.projekt,
        )

        self.projekt.aufwand = self.aufwand
        self.haushaltsposten.projekte = [self.projekt]
        self.haushaltsplan.haushaltsposten_liste = [self.haushaltsposten]

    def test_generate_excel(self):
        # Instantiate the HaushaltsplanExcelGenerator with the test Haushaltsplan
        generator = HaushaltsplanExcelGenerator(haushaltsplan=self.haushaltsplan)

        # Generate the Excel file
        file_name = 'test_haushaltsplan.xlsx'
        generator.generate_excel(file_name=file_name)

        # Check if the file was generated successfully
        self.assertTrue(os.path.exists(file_name))

        # Open the generated Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active

        # Check the headers in the first row
        headers = ['Projekt Name', 'Einnahmen', 'Ausgaben']
        for col, header in enumerate(headers, start=1):
            cell_value = worksheet.cell(row=1, column=col).value
            self.assertEqual(cell_value, header)

        # Check the Haushaltsplan data in rows 3 to 7
        data = [
            ('Name:', 'Test Plan'),
            ('Standort:', 'Heilbronn'),
            ('Ersteller:', 'testuser'),
            ('Startjahr:', 2023),
            ('Studierendenanzahl:', 1000),
        ]
        for row, (label, value) in enumerate(data, start=3):
            self.assertEqual(worksheet.cell(row=row, column=1).value, label)
            self.assertEqual(worksheet.cell(row=row, column=2).value, value)

        # Check the Haushaltsposten data in row 9
        self.assertEqual(worksheet.cell(row=9, column=1).value, 'Test Posten')

        # Check the Projekt data in row 10
        self.assertEqual(worksheet.cell(row=10, column=1).value, 'Test Projekt')
        self.assertEqual(worksheet.cell(row=10, column=2).value, 5000)
        self.assertEqual(worksheet.cell(row=10, column=3).value, 3000)

        # Close the workbook and delete the generated file
        workbook.close()
        os.remove(file_name)
