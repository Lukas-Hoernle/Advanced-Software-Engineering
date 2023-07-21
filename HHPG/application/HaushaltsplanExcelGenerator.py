from datetime import datetime

import openpyxl

from HHPG.domain.entity import haushaltsplan


class HaushaltsplanExcelGenerator:

    def __init__(self, haushaltsplan: haushaltsplan):
        self.haushaltsplan = haushaltsplan

    def generate_excel(self, file_name=None):
        if file_name is None:
            file_name = datetime.now().strftime('%Y-%m-%d_%H-%M-%S.xls')

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        builder = ExcelBuilder(worksheet)

        builder.set_headers(
            'Projekt Name',
            'Einnahmen',
            'Ausgaben'
        )
        builder.write_headers()

        row_index = 3

        builder.set_haushaltsplan(
            self.haushaltsplan,
            row_index
        )
        builder.write_haushaltsplan()

        row_index = 9

        haushaltsposten_list = self.haushaltsplan.haushaltsposten_liste

        for haushaltsposten in haushaltsposten_list:
            builder.set_haushaltsposten(
                haushaltsposten,
                row_index
            )
            builder.write_haushaltsposten()
            row_index += 1

            projekte_list = haushaltsposten.projekte

            for projekt in projekte_list:
                builder.set_projekt(
                    projekt,
                    row_index
                )
                builder.write_projekt()
                row_index += 1

        workbook.save(file_name)


class ExcelBuilder:

    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.headers = []
        self.haushaltsposten = None
        self.projekt = None
        self.row_index = None

    def set_headers(self, *headers):
        self.headers = headers

    def write_headers(self):
        for col, header in enumerate(
                self.headers,
                start=1
        ):
            self.worksheet.cell(
                row=1,
                column=col,
                value=header
            )

    def set_haushaltsplan(
            self,
            haushaltsplan,
            row_index
    ):
        self.haushaltsplan = haushaltsplan
        self.row_index = row_index

    def write_haushaltsplan(self):
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value="Name:"
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.haushaltsplan.plan_name
        )
        self.row_index += 1
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value="Standort:"
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.haushaltsplan.standort
        )
        self.row_index += 1
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value="Ersteller:"
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.haushaltsplan.author.username
        )
        self.row_index += 1
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value="Startjahr:"
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.haushaltsplan.startjahr
        )
        self.row_index += 1
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value="Studierendenanzahl:"
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.haushaltsplan.studierendenzahl
        )

    def set_haushaltsposten(
            self,
            haushaltsposten,
            row_index,
    ):
        self.haushaltsposten = haushaltsposten
        self.row_index = row_index

    def write_haushaltsposten(self):
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value=self.haushaltsposten.posten_name
        )

    def set_projekt(
            self,
            projekt,
            row_index
    ):
        self.projekt = projekt
        self.row_index = row_index

    def write_projekt(self):
        self.worksheet.cell(
            row=self.row_index,
            column=1,
            value=self.projekt.projekt_name
        )
        self.worksheet.cell(
            row=self.row_index,
            column=2,
            value=self.projekt.aufwand.einnahmen
        )
        self.worksheet.cell(
            row=self.row_index,
            column=3,
            value=self.projekt.aufwand.ausgaben
        )
